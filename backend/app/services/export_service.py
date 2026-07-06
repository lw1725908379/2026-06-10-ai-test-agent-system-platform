"""
测试用例导出服务

提供 BDD 测试用例导出为 .feature 文件、以及 Excel 导出功能。
"""

import io
import zipfile
from datetime import datetime
from typing import Tuple
from uuid import uuid4

from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.test_case import TestCase
from app.repositories.test_case_repo import TestCaseRepository
from app.repositories.project_repo import ProjectRepository
from app.schemas.enums import ExportStatus, TestCaseTemplate
from app.schemas.test_case import (
    ExportBDDRequest, ExportBDDResponse, ExportStatusResponse
)
from app.utils.exceptions import NotFoundException, BadRequestException
from app.config.settings import settings


_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_EXCEL_COLUMN_WIDTHS = {
    "A": 18,
    "B": 35,
    "C": 14,
    "D": 12,
    "E": 10,
    "F": 30,
    "G": 40,
    "H": 30,
    "I": 40,
    "J": 20,
}


class ExportService:
    """
    测试用例导出服务

    处理 BDD 导出为 .feature / .zip，以及 Excel 导出的逻辑。
    """

    COLLECTION_NAME = "export_jobs"

    def __init__(self, db: AsyncSession, mongodb: AsyncIOMotorDatabase):
        self.db = db
        self.mongodb = mongodb
        self.tc_repo = TestCaseRepository(db)
        self.project_repo = ProjectRepository(db)

    # ========================================================================
    # BDD 导出
    # ========================================================================

    async def start_bdd_export(
        self,
        project_identifier: str,
        data: ExportBDDRequest
    ) -> ExportBDDResponse:
        """
        启动 BDD 测试用例导出任务
        """
        if not data.test_case_ids:
            raise BadRequestException("请选择至少一个要导出的测试用例")

        export_id = str(uuid4())

        export_job = {
            "_id": export_id,
            "project_identifier": project_identifier,
            "test_case_ids": data.test_case_ids,
            "combine_into_one": data.combine_into_one,
            "combined_feature": data.combined_feature,
            "combined_background": data.combined_background,
            "status": ExportStatus.PENDING.value,
            "download_url": None,
            "file_content": None,
            "filename": None,
            "content_type": None,
            "error_message": None,
            "created_at": datetime.utcnow(),
            "completed_at": None,
        }

        await self.mongodb[self.COLLECTION_NAME].insert_one(export_job)

        # 异步处理导出任务（当前为同步处理，后续可改为后台队列）
        await self._process_bdd_export(export_id)

        status_url = f"{settings.api_prefix}/exports/{export_id}/status"

        return ExportBDDResponse(
            success=True,
            export_id=export_id,
            status=ExportStatus.PENDING,
            status_url=status_url
        )

    async def _process_bdd_export(self, export_id: str) -> None:
        """处理 BDD 导出任务"""
        try:
            await self.mongodb[self.COLLECTION_NAME].update_one(
                {"_id": export_id},
                {"$set": {"status": ExportStatus.PROCESSING.value}}
            )

            job = await self.mongodb[self.COLLECTION_NAME].find_one({"_id": export_id})
            if not job:
                return

            project = await self.project_repo.get_by_identifier(job["project_identifier"])
            if not project:
                raise NotFoundException("项目", job["project_identifier"])

            test_cases = await self.tc_repo.get_by_identifiers(
                project.id, job["test_case_ids"]
            )
            if not test_cases:
                raise NotFoundException("测试用例", str(job["test_case_ids"]))

            # 过滤出 BDD 用例；普通用例跳过并记录提示
            bdd_cases = [tc for tc in test_cases if tc.template == TestCaseTemplate.TEST_CASE_BDD]
            if not bdd_cases:
                raise BadRequestException("选中的测试用例中没有 BDD 模板用例，无法导出 .feature")

            if job["combine_into_one"]:
                filename = f"{job['combined_feature'].replace(' ', '_')}.feature"
                content_type = "text/plain"
                file_content = self._generate_combined_feature(
                    job["combined_feature"],
                    job.get("combined_background"),
                    bdd_cases,
                ).encode("utf-8")
            else:
                filename = f"bdd_export_{export_id[:8]}.zip"
                content_type = "application/zip"
                file_content = self._create_bdd_zip(bdd_cases)

            download_url = f"{settings.api_prefix}/exports/{export_id}/download"

            await self.mongodb[self.COLLECTION_NAME].update_one(
                {"_id": export_id},
                {
                    "$set": {
                        "status": ExportStatus.COMPLETED.value,
                        "download_url": download_url,
                        "file_content": file_content,
                        "filename": filename,
                        "content_type": content_type,
                        "completed_at": datetime.utcnow(),
                    }
                }
            )

        except Exception as e:
            await self.mongodb[self.COLLECTION_NAME].update_one(
                {"_id": export_id},
                {
                    "$set": {
                        "status": ExportStatus.FAILED.value,
                        "error_message": str(e),
                        "completed_at": datetime.utcnow(),
                    }
                }
            )

    def _generate_combined_feature(
        self,
        feature_name: str,
        background: str | None,
        test_cases: list[TestCase],
    ) -> str:
        """生成合并的 .feature 文件内容"""
        lines = [f"Feature: {feature_name}"]

        if background:
            lines.extend(["", "  Background:", f"    {background}"])

        for tc in test_cases:
            scenario_lines = self._normalize_scenario_lines(tc.scenario)
            if scenario_lines:
                lines.append("")
                for line in scenario_lines:
                    lines.append(f"  {line}")
            else:
                lines.extend(["", f"  Scenario: {tc.name or tc.identifier}"])
            if tc.background:
                lines.append(f"    {tc.background}")
            if not scenario_lines:
                # 兜底：从 preconditions / description 构造
                if tc.preconditions:
                    lines.append(f"    Given {tc.preconditions}")
                if tc.description:
                    lines.append(f"    When {tc.description}")
                    lines.append(f"    Then 验证{tc.name or '结果'}")

        return "\n".join(lines)

    def _normalize_scenario_lines(self, scenario: str | None) -> list[str]:
        """
        规范化 scenario 字段内容。

        如果 scenario 已包含 "Scenario:" 标题行，则直接拆分为多行；
        否则返回空列表，由调用方自行添加 Scenario 标题。
        """
        if not scenario:
            return []
        lines = [line.strip() for line in scenario.splitlines() if line.strip()]
        if lines and lines[0].lower().startswith("scenario:"):
            return lines
        return []

    def _generate_single_feature(self, test_case: TestCase) -> str:
        """为单个 BDD 测试用例生成 .feature 内容"""
        feature = test_case.feature or f"Feature: {test_case.name or test_case.identifier}"
        lines = [feature, ""]

        if test_case.background:
            lines.append("  Background:")
            for line in test_case.background.splitlines():
                lines.append(f"    {line.strip()}")
            lines.append("")

        scenario_lines = self._normalize_scenario_lines(test_case.scenario)
        if scenario_lines:
            for line in scenario_lines:
                lines.append(f"  {line}")
        else:
            lines.append(f"  Scenario: {test_case.name or test_case.identifier}")
        if not scenario_lines and test_case.scenario:
            if test_case.preconditions:
                lines.append(f"    Given {test_case.preconditions}")
            if test_case.description:
                lines.append(f"    When {test_case.description}")
                lines.append(f"    Then 验证{test_case.name or '结果'}")

        return "\n".join(lines)

    def _create_bdd_zip(self, test_cases: list[TestCase]) -> bytes:
        """为每个 BDD 用例生成 .feature 并打包为 zip"""
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for tc in test_cases:
                filename = f"{tc.identifier}.feature"
                content = self._generate_single_feature(tc)
                zf.writestr(filename, content)
        buffer.seek(0)
        return buffer.read()

    # ========================================================================
    # Excel 导出
    # ========================================================================

    async def export_test_cases_to_excel(
        self,
        project_identifier: str,
        test_case_ids: list[str],
    ) -> Tuple[bytes, str]:
        """
        将选中的测试用例导出为 Excel 文件。

        Returns:
            Tuple[bytes, str]: (文件内容, 文件名)
        """
        if not test_case_ids:
            raise BadRequestException("请选择至少一个要导出的测试用例")

        project = await self.project_repo.get_by_identifier(project_identifier)
        if not project:
            raise NotFoundException("项目", project_identifier)

        test_cases = await self.tc_repo.get_by_identifiers(project.id, test_case_ids)
        if not test_cases:
            raise NotFoundException("测试用例", str(test_case_ids))

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("无法创建工作表")
        ws.title = "测试用例"

        headers = [
            "用例编号",
            "用例标题",
            "所属模块",
            "用例类型",
            "优先级",
            "前置条件",
            "测试步骤",
            "测试数据",
            "预期结果",
            "备注",
        ]
        ws.append(headers)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _ALIGNMENT_CENTER
            cell.border = _BORDER

        for tc in test_cases:
            if tc.template == TestCaseTemplate.TEST_CASE_BDD:
                steps_text = tc.scenario or ""
                expected_text = ""
                module_text = tc.feature or ""
            else:
                steps_lines = []
                expected_lines = []
                for idx, step in enumerate(tc.steps or [], start=1):
                    action = step.action or ""
                    expected = step.expected_result or ""
                    steps_lines.append(f"{idx}. {action}")
                    expected_lines.append(f"{idx}. {expected}")
                steps_text = "\n".join(steps_lines)
                expected_text = "\n".join(expected_lines)
                module_text = ""

            row = [
                tc.identifier,
                tc.name,
                module_text,
                "BDD" if tc.template == TestCaseTemplate.TEST_CASE_BDD else "功能",
                tc.priority.value if tc.priority else "",
                tc.preconditions or "",
                steps_text,
                "",
                expected_text,
                tc.description or "",
            ]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = _ALIGNMENT_WRAP
                cell.border = _BORDER

        for col_letter, width in _EXCEL_COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[1].height = 24
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 60

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"test_cases_{project_identifier}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return buffer.read(), filename

    # ========================================================================
    # 任务状态 / 下载
    # ========================================================================

    async def get_export_status(self, export_id: str) -> ExportStatusResponse:
        """获取导出任务状态"""
        job = await self.mongodb[self.COLLECTION_NAME].find_one({"_id": export_id})
        if not job:
            raise NotFoundException(f"导出任务 {export_id} 不存在")

        return ExportStatusResponse(
            success=True,
            export_id=export_id,
            status=ExportStatus(job["status"]),
            download_url=job.get("download_url"),
            error_message=job.get("error_message")
        )

    async def download_export(
        self,
        export_id: str
    ) -> Tuple[bytes, str, str]:
        """下载导出文件"""
        job = await self.mongodb[self.COLLECTION_NAME].find_one({"_id": export_id})
        if not job:
            raise NotFoundException(f"导出任务 {export_id} 不存在")

        if job["status"] != ExportStatus.COMPLETED.value:
            raise BadRequestException(f"导出任务尚未完成，当前状态: {job['status']}")

        return (
            job["file_content"],
            job["filename"],
            job["content_type"]
        )
