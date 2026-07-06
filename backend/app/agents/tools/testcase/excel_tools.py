"""
测试用例 Excel 导出工具

提供将测试用例导出为 Excel 文件的能力，支持企业级测试管理工具的导入格式。
同时兼容后端 API 返回的测试用例数据结构。
"""

import io
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from langchain.tools import tool

logger = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# fmt: off  MC80OmFIVnBZMlhsaUpqbWxvYzZaMjVsVnc9PTpjOWJlNmRjMA==

_DEFAULT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 35,
    "C": 14,
    "D": 12,
    "E": 10,
    "F": 30,
    "G": 40,
    "H": 30,
    "I": 40,
    "J": 20,
}


def _flatten_steps(steps: list[dict[str, Any]] | list[str] | str | None) -> tuple[str, str]:
    """
    将步骤列表转换为带序号的文本，并尝试提取预期结果。

    兼容三种输入格式：
      - 字符串：直接作为步骤返回，预期结果为空
      - 字符串列表：按序号拼接为步骤
      - 字典列表：支持以下键
          * 后端 API 格式: action / expected_result
          * 通用格式      : step / action / 操作描述 / data / target

    Returns:
        (steps_text, expected_results_text)
    """
    if not steps:
        return "", ""
    if isinstance(steps, str):
        return steps, ""

    step_lines = []
    expected_lines = []
    for idx, step in enumerate(steps, start=1):
        if isinstance(step, str):
            step_lines.append(f"{idx}. {step}")
            continue

        # 后端 API 格式: {action, expected_result}
        action = step.get("action", "")
        expected = step.get("expected_result", "")
        if action:
            step_lines.append(f"{idx}. {action}")
        if expected:
            expected_lines.append(f"{idx}. {expected}")

        # 如果 action 为空，尝试通用格式
        if not action:
            seq = step.get("seq", step.get("step", idx))
            action_text = step.get("action", step.get("操作描述", ""))
            target = step.get("target", step.get("操作对象", ""))
            data = step.get("data", "")
            line = f"{seq}. {action_text}"
            if target:
                line += f" [{target}]"
            if data:
                line += f"（数据：{data}）"
            if action_text or target or data:
                step_lines.append(line)

    return "\n".join(step_lines), "\n".join(expected_lines)


def _flatten_test_data(test_data: dict[str, Any] | str | None) -> str:
    """将测试数据转换为文本。"""
    if not test_data:
        return ""
    if isinstance(test_data, str):
        return test_data
    lines = [f"{k}: {v}" for k, v in test_data.items()]
    return "\n".join(lines)


def _flatten_expected_results(expected_results: list[str] | str | None) -> str:
    """将预期结果列表转换为文本。"""
    if not expected_results:
        return ""
    if isinstance(expected_results, str):
        return expected_results
    lines = []
    for idx, result in enumerate(expected_results, start=1):
        lines.append(f"{idx}. {result}")
    return "\n".join(lines)


def _flatten_preconditions(preconditions: list[str] | str | None) -> str:
    """将前置条件转换为文本。"""
    if not preconditions:
        return ""
    if isinstance(preconditions, str):
        return preconditions
    lines = []
    for idx, cond in enumerate(preconditions, start=1):
        lines.append(f"{idx}. {cond}")
    return "\n".join(lines)
# pylint: disable  MS80OmFIVnBZMlhsaUpqbWxvYzZaMjVsVnc9PTpjOWJlNmRjMA==

# pragma: no cover  Mi80OmFIVnBZMlhsaUpqbWxvYzZaMjVsVnc9PTpjOWJlNmRjMA==


def _extract_field(case: dict[str, Any], *keys: str, default: Any = "") -> Any:
    """从字典中按多个候选键提取值。"""
    for key in keys:
        if key in case:
            return case[key]
    return default


def _normalize_case(case: dict[str, Any]) -> dict[str, Any]:
    """将后端 API 格式或多种手写格式统一为 Excel 工具内部格式。"""
    # 步骤与预期结果
    raw_steps = _extract_field(
        case,
        "steps", "测试步骤", "test_case_steps", "step_list",
        default=None,
    )
    steps_text, extracted_expected = _flatten_steps(raw_steps)

    # 预期结果：优先使用独立字段，其次从步骤中提取
    expected_results = _extract_field(
        case,
        "expected_results", "预期结果", "expected", "expected_result",
        default=None,
    )
    if not expected_results and extracted_expected:
        expected_results = extracted_expected

    # 备注：合并 remarks / description / tags
    remarks = _extract_field(case, "remarks", "备注", default="")
    description = _extract_field(case, "description", default="")
    tags = case.get("tags", [])
    extra_parts = []
    if description:
        extra_parts.append(f"描述：{description}")
    if tags:
        if isinstance(tags, list):
            extra_parts.append(f"标签：{', '.join(str(t) for t in tags)}")
        else:
            extra_parts.append(f"标签：{tags}")
    if extra_parts:
        remarks = "\n".join([remarks] + extra_parts) if remarks else "\n".join(extra_parts)

    return {
        "id": _extract_field(case, "id", "identifier", "用例编号"),
        "title": _extract_field(case, "title", "name", "用例标题"),
        "module": _extract_field(case, "module", "所属模块"),
        "type": _extract_field(case, "type", "case_type", "用例类型"),
        "priority": _extract_field(case, "priority", "优先级"),
        "preconditions": _flatten_preconditions(
            _extract_field(case, "preconditions", "前置条件", default=None)
        ),
        "steps": steps_text,
        "test_data": _flatten_test_data(
            _extract_field(case, "test_data", "测试数据", "data", default=None)
        ),
        "expected_results": _flatten_expected_results(expected_results),
        "remarks": remarks,
    }


@tool
def export_test_cases_to_excel(
    test_cases: list[dict[str, Any]],
    output_path: str | Path | None = None,
    sheet_name: str = "测试用例",
) -> str:
    """
    将测试用例列表导出为 Excel 文件。

    当用户说"导出Excel"、"导出为excel表格"、"生成Excel"等时，必须调用此工具，
    而不是返回 JSON 或 Markdown 表格。

    支持的测试用例字段（兼容后端 API 与多种手写格式）：
      - id / identifier / 用例编号
      - title / name / 用例标题
      - module / 所属模块
      - type / case_type / 用例类型
      - priority / 优先级
      - preconditions / 前置条件
      - steps / 测试步骤 / test_case_steps
      - test_data / 测试数据
      - expected_results / 预期结果 / expected_result
      - remarks / 备注 / description / tags

    Args:
        test_cases: 测试用例字典列表，每个字典描述一条用例。
        output_path: 导出的 Excel 文件路径（支持 str 或 Path）。
            默认为 /app/backend/workspace/testcase/exports/测试用例_<时间戳>.xlsx
        sheet_name: 工作表名称，默认为 "测试用例"。

    Returns:
        导出文件的绝对路径字符串。
    """
    if not test_cases:
        raise ValueError("测试用例列表为空，无法导出 Excel。")

    if output_path is None:
        default_dir = Path("/app/backend/workspace/testcase/exports")
        default_dir.mkdir(parents=True, exist_ok=True)
        output_path = default_dir / f"测试用例_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表。")
    ws.title = sheet_name

    headers = [
        "用例编号",
        "用例标题",
        "所属模块",
        "用例类型",
        "优先级",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
        "备注",
    ]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGNMENT_CENTER
        cell.border = _BORDER

    for case in test_cases:
        normalized = _normalize_case(case)
        row = [
            normalized["id"],
            normalized["title"],
            normalized["module"],
            normalized["type"],
            normalized["priority"],
            normalized["preconditions"],
            normalized["steps"],
            normalized["test_data"],
            normalized["expected_results"],
            normalized["remarks"],
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER

    for col_letter, width in _DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60
# pragma: no cover  My80OmFIVnBZMlhsaUpqbWxvYzZaMjVsVnc9PTpjOWJlNmRjMA==

    filename = output_path.name

    from app.config.minio_client import MinIOClient

    # 保存到内存中
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 上传到 MinIO
    object_name = f"exports/{filename}"
    try:
        MinIOClient.upload_bytes(
            object_name=object_name,
            data=buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # 下载通过后端 API 代理（从 MinIO 读取后下发）
        download_url = f"http://localhost:3000/api/v2/exports/download/{filename}"
    except Exception as e:
        # MinIO 上传失败时回退到本地文件
        wb.save(str(output_path))
        download_url = f"http://localhost:3000/api/v2/exports/download/{filename}"
        logger.warning(f"MinIO 上传失败，回退到本地文件: {e}")

    return json.dumps({
        "success": True,
        "filename": filename,
        "download_url": download_url,
        "message": f"测试用例已导出为 {filename}，点击以下链接下载：\n{download_url}"
    }, ensure_ascii=False)
